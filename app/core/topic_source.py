"""
Excel-based topic source. Replaces Gemini topic generation.
Reads topics from an .xlsx file and tracks consumption in the DB
so the original spreadsheet is never modified.
"""

import logging
import os

import openpyxl

from db.db import Database

logger = logging.getLogger(__name__)


class TopicSource:
    """Reads topics from an Excel file and tracks consumption in DB."""

    def __init__(self, db: Database, config: dict) -> None:
        """
        Args:
            db:     Database instance (thread-safe).
            config: Full prompt_templates config dict.
        """
        self._db = db
        self._config = config

    def _get_required_columns(self) -> list[str]:
        """Return the required column names from config."""
        ts_cfg = self._config.get("topic_source", {})
        return ts_cfg.get(
            "required_columns",
            ["Topic", "Subject", "Headline", "Fact_Text",
             "Highlight_1", "Highlight_2", "Category"],
        )

    def validate_file(self, file_path: str) -> tuple[bool, str]:
        """
        Validate that file_path is a readable .xlsx with the required columns.
        Returns (ok, message).
        """
        if not file_path or not os.path.isfile(file_path):
            return False, f"File not found: {file_path}"

        if not file_path.lower().endswith(".xlsx"):
            return False, "File must be an .xlsx Excel file."

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                wb.close()
                return False, "Excel file has no active worksheet."

            # Read header row
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            wb.close()

            if header_row is None:
                return False, "Excel file is empty (no header row)."

            headers = [str(h).strip() if h else "" for h in header_row]
            required = self._get_required_columns()
            missing = [col for col in required if col not in headers]

            if missing:
                return False, (
                    f"Missing required columns: {', '.join(missing)}. "
                    f"Found: {', '.join(headers)}. "
                    f"Required: {', '.join(required)}"
                )

            return True, "File validated successfully."

        except Exception as e:
            return False, f"Could not read Excel file: {e}"

    def get_next_topic(self, active_file_path: str) -> dict | None:
        """
        Read the xlsx, find the first row not yet consumed in topic_source_rows.

        If found:
          1. Insert a content_items row pre-filled with the Excel fields.
          2. Insert a topic_source_rows row marking the row consumed.
          3. Return the new content_item dict.

        If no unconsumed rows remain, return None.
        """
        rows = self._load_rows(active_file_path)
        if not rows:
            return None

        consumed_indices = self._db.get_consumed_row_indices(active_file_path)

        for row_data in rows:
            idx = row_data["_row_index"]
            if idx in consumed_indices:
                continue

            # Found an unconsumed row — create the content item
            topic = row_data.get("Topic", "")
            if not topic:
                # Skip rows with empty topic
                continue

            item_id = self._db.create_content_item(
                topic=topic,
                subject=row_data.get("Subject", ""),
                headline=row_data.get("Headline", ""),
                fact_text=row_data.get("Fact_Text", ""),
                highlight_1=row_data.get("Highlight_1", ""),
                highlight_2=row_data.get("Highlight_2", ""),
                category=row_data.get("Category", ""),
            )

            self._db.mark_row_consumed(
                source_file=active_file_path,
                row_index=idx,
                topic=topic,
                content_item_id=item_id,
            )

            item = self._db.get_content_item(item_id)
            logger.info(
                "Pulled topic from row %d of '%s': '%s' → item #%d",
                idx, os.path.basename(active_file_path), topic[:60], item_id,
            )
            return item

        # All rows consumed
        return None

    def get_next_topics(self, active_file_path: str, count: int = 3) -> list[dict]:
        """Pull up to `count` unconsumed topics from the file."""
        items = []
        for _ in range(count):
            item = self.get_next_topic(active_file_path)
            if item is None:
                break
            items.append(item)
        return items

    def get_remaining_count(self, active_file_path: str) -> int:
        """Return the number of unconsumed rows for the given file."""
        rows = self._load_rows(active_file_path)
        if not rows:
            return 0
        consumed_indices = self._db.get_consumed_row_indices(active_file_path)
        return sum(1 for r in rows if r["_row_index"] not in consumed_indices)

    def _load_rows(self, file_path: str) -> list[dict]:
        """
        Load all data rows from the xlsx, mapping column headers to values.
        Each dict includes a '_row_index' key (0-based index among data rows).
        """
        if not file_path or not os.path.isfile(file_path):
            logger.warning("Topic source file not found: %s", file_path)
            return []

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                wb.close()
                return []

            all_rows = list(ws.iter_rows(values_only=True))
            wb.close()

            if len(all_rows) < 2:
                return []

            headers = [str(h).strip() if h else "" for h in all_rows[0]]
            result = []
            for i, row_vals in enumerate(all_rows[1:]):
                row_dict = {"_row_index": i}
                for col_idx, header in enumerate(headers):
                    if header and col_idx < len(row_vals):
                        val = row_vals[col_idx]
                        row_dict[header] = str(val).strip() if val is not None else ""
                    elif header:
                        row_dict[header] = ""
                result.append(row_dict)

            return result

        except Exception as e:
            logger.error("Failed to load rows from '%s': %s", file_path, e)
            return []
